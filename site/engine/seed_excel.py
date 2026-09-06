"""Загрузка плана и факта из Excel FCF в SQLite."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from .categories import EXPENSE_CATEGORIES, INCOME_CATEGORIES
from .db import set_meta, upsert_ledger

EXCEL_CANDIDATES = [
    Path("/Users/Sedova.Maria/Desktop/Саша/Мониторинг бюджета/5.09.2026_Бюджет 2026.xlsx"),
    Path("/Users/Sedova.Maria/Desktop/Саша/Мониторинг бюджета/16.08.2026_Бюджет 2026.xlsx"),
    Path("/Users/Sedova.Maria/Desktop/Саша/Мониторинг бюджета/16.07.2026_Бюджет 2026.xlsx"),
]

PLAN_INCOME = {
    "Зарплата Саша": 5,
    "Премия Саша": 6,
    "Продажа квартиры Саша": 7,
    "Зарплата Маша": 8,
    "Премия Маша": 9,
    "Займы": 10,
    "Подарки": 11,
}
PLAN_EXPENSE = {
    "Ипотека платеж": 20,
    "Дедушка долг": 21,
    "Ремонт квартиры": 22,  # в Excel: «Ремонт в квартире»
    "Квартира Тайланд": 23,
    "Свадебное путешествие": 24,
    "Саша учеба": 25,
    "Парковка": 26,
    "Отпуска": 27,
    "Страховка": 28,
    "Налоги": 29,
    "Ребенок": 30,
    "Супермаркеты": 31,
    "Такси": 32,
    "Рестораны": 33,
    "Одежда и обувь": 34,
    "Квартплата": 35,
    "Мобильная связь": 36,
    "Товары для дома": 37,
    "Косметика": 38,
    "Развлечения": 39,
    "Бьюти процедуры": 40,
    "Парковки и штрафы": 41,
    "Бензин": 42,
    "Переводы": 43,
    "Прочее": 44,
    "Расходы на семьи": 45,
    "Подарки друг другу": 46,
    "Крупные покупки": 47,
    "Абонемент в спорт-зал": 48,
}
FACT_INCOME = PLAN_INCOME
FACT_EXPENSE = {
    "Ипотека платеж": 25,
    "Дедушка долг": 26,
    "Ремонт квартиры": 27,
    "Квартира Тайланд": 28,
    "Свадебное путешествие": 29,
    "Саша учеба": 30,
    "Парковка": 31,
    "Отпуска": 32,
    "Страховка": 33,
    "Налоги": 34,
    "Ребенок": 35,
    "Супермаркеты": 36,
    "Такси": 37,
    "Рестораны": 38,
    "Одежда и обувь": 39,
    "Квартплата": 40,
    "Мобильная связь": 41,
    "Товары для дома": 42,
    "Косметика": 43,
    "Развлечения": 44,
    "Бьюти процедуры": 45,
    "Парковки и штрафы": 46,
    "Бензин": 47,
    "Переводы": 48,
    "Прочее": 49,
    "Расходы на семьи": 50,
    "Подарки друг другу": 51,
    "Крупные покупки": 52,
    "Абонемент в спорт-зал": 53,
}


def _num(value) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def find_excel() -> Path | None:
    for path in EXCEL_CANDIDATES:
        if path.exists():
            return path
    root = Path("/Users/Sedova.Maria/Desktop/Саша/Мониторинг бюджета")
    files = sorted(root.glob("*Бюджет 2026.xlsx"), reverse=True)
    return files[0] if files else None


def seed_from_excel(conn, path: Path | None = None) -> str:
    excel = path or find_excel()
    if not excel:
        raise FileNotFoundError("Не найден файл бюджета Excel")

    wb = openpyxl.load_workbook(excel, data_only=True)
    plan = wb["FCF 2026 ПЛАН"]
    fact = wb["FCF 2026 ФАКТ"]

    # План: колонки по годам 2026–2040 (по 12 месяцев), как в Excel
    for year in range(2026, 2041):
        for month in range(1, 13):
            col = 3 + (year - 2026) * 12 + (month - 1)
            for cat, row in PLAN_INCOME.items():
                upsert_ledger(
                    conn, year, month, cat, "income",
                    plan=_num(plan.cell(row, col).value), source="excel",
                )
            for cat, row in PLAN_EXPENSE.items():
                upsert_ledger(
                    conn, year, month, cat, "expense",
                    plan=_num(plan.cell(row, col).value), source="excel",
                )

    # Факт только 2026
    year = 2026
    for month in range(1, 13):
        col = 2 + month  # C=3 for January
        for cat, row in FACT_INCOME.items():
            upsert_ledger(
                conn, year, month, cat, "income",
                fact=_num(fact.cell(row, col).value), source="excel",
            )
        for cat, row in FACT_EXPENSE.items():
            upsert_ledger(
                conn, year, month, cat, "expense",
                fact=_num(fact.cell(row, col).value), source="excel",
            )

        # Август закрыт в выгрузке 5.09; сентябрь — частичный; окт–дек — ещё план
        if month == 9:
            conn.execute(
                "UPDATE ledger SET source='partial' WHERE year=? AND month=?",
                (year, month),
            )
        elif month >= 10:
            conn.execute(
                "UPDATE ledger SET source='forecast' WHERE year=? AND month=?",
                (year, month),
            )

    set_meta(conn, "excel_file", excel.name)
    set_meta(conn, "seeded", "1")
    set_meta(conn, "closed_month", "8")
    conn.commit()
    return excel.name
