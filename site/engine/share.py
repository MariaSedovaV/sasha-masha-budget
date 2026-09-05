"""Стоимость активов: доли + timeline из ledger (не Excel «Разделение_деньги»)."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from .assets import build_asset_timeline
from .markets import fetch_markets
from .seed_excel import find_excel

FALLBACK_PROPERTY = [
    {"name": "Квартира Тайланд", "owner": "Маша", "share": 0.5, "comment": None},
    {"name": "Квартира Тайланд", "owner": "Саша", "share": 0.5, "comment": None},
    {"name": "Квартира Петербург", "owner": "Саша", "share": 1.0, "comment": None},
    {"name": "Парковочное место", "owner": "Маша", "share": 1.0, "comment": None},
]


def _num(value) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def read_property_shares(excel: Path | None = None) -> list[dict]:
    path = excel or find_excel()
    props = list(FALLBACK_PROPERTY)
    if not path or not path.exists():
        return props
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return props
    if "Разделение_недвижимость" not in wb.sheetnames:
        return props
    ws = wb["Разделение_недвижимость"]
    out = []
    for r in range(2, (ws.max_row or 1) + 1):
        name = ws.cell(r, 1).value
        owner = ws.cell(r, 2).value
        share = _num(ws.cell(r, 3).value)
        comment = ws.cell(r, 4).value
        if name and owner:
            out.append(
                {
                    "name": str(name).strip(),
                    "owner": str(owner).strip(),
                    "share": share,
                    "comment": str(comment) if comment else None,
                }
            )
    return out or props


def build_share(rows: list[dict], closed_month: int = 7) -> dict:
    try:
        markets = fetch_markets()
    except Exception:
        markets = {}

    timeline = build_asset_timeline(rows, closed_month, markets)
    cur = timeline["current"]
    liq = timeline["liquid"]

    return {
        "source": "ledger + рыночные ориентиры (Куинджи / Bangtao)",
        "totals": {
            "cash": cur["cash"],
            "masha_cash": cur["masha"],
            "sasha_cash": cur["sasha"],
            "gold": cur["gold"],
            "spb": cur["spb"],
            "phuket": cur["phuket"],
            "liquid": liq["liquid_total"],
            "all": timeline["kpis"]["now"],
        },
        "property": timeline["property_shares"],
        "timeline": timeline,
        "markets": {
            "usd": (markets.get("usd") or {}).get("value"),
            "thb": (markets.get("thb") or {}).get("value"),
            "gold_gram": (markets.get("gold_gram") or {}).get("value"),
            "as_of": markets.get("as_of"),
        },
    }
