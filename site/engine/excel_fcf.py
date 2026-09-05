"""Чтение кумулятива FCF и накоплений напрямую из Excel (как в файле)."""

from __future__ import annotations

from functools import lru_cache
from pathlib import Path

import openpyxl

from .seed_excel import find_excel, _num

BASE_YEAR = 2026
PLAN_END_YEAR = 2028
PLAN_SHEET = "FCF 2026 ПЛАН"
FACT_SHEET = "FCF 2026 ФАКТ"

PLAN_CUMUL_ROW = 50
PLAN_TOTAL_ROW = 48
FACT_CUMUL_ROW = 55
FACT_TOTAL_ROW = 53
# Балансы и месячные потоки накоплений на ФАКТ
FACT_MASHA_BAL_ROW = 13
FACT_SASHA_BAL_ROW = 14
FACT_MASHA_FLOW_ROW = 18
FACT_SASHA_FLOW_ROW = 19


def _plan_col(year: int, month: int) -> int:
    return 3 + (year - BASE_YEAR) * 12 + (month - 1)


def _fact_col(month: int) -> int:
    return 2 + month


@lru_cache(maxsize=4)
def _workbook(path_str: str):
    return openpyxl.load_workbook(path_str, data_only=True)


def _wb(path: Path | None = None):
    excel = path or find_excel()
    if not excel or not excel.exists():
        return None, None
    try:
        return _workbook(str(excel.resolve())), excel
    except Exception:
        return None, excel


def read_plan_horizon(path: Path | None = None) -> dict | None:
    """План: кумулятив и месячный ИТОГО за 2026–2028 из листа FCF ПЛАН."""
    wb, excel = _wb(path)
    if not wb or PLAN_SHEET not in wb.sheetnames:
        return None
    ws = wb[PLAN_SHEET]
    labels, cumul, monthly = [], [], []
    for year in range(BASE_YEAR, PLAN_END_YEAR + 1):
        for month in range(1, 13):
            col = _plan_col(year, month)
            labels.append(f"{month:02d}.{year}")
            cumul.append(_num(ws.cell(PLAN_CUMUL_ROW, col).value))
            monthly.append(_num(ws.cell(PLAN_TOTAL_ROW, col).value))
    return {
        "labels": labels,
        "cumul": cumul,
        "monthly": monthly,
        "excel": excel.name if excel else None,
        "end_2027": cumul[23] if len(cumul) > 23 else None,
        "end_2028": cumul[-1] if cumul else None,
    }


def read_fact_cumul_series(path: Path | None = None) -> dict | None:
    """Факт: кумулятив и ИТОГО по месяцам 2026 из листа FCF ФАКТ."""
    wb, excel = _wb(path)
    if not wb or FACT_SHEET not in wb.sheetnames:
        return None
    ws = wb[FACT_SHEET]
    cumul, monthly = [], []
    for month in range(1, 13):
        col = _fact_col(month)
        cumul.append(_num(ws.cell(FACT_CUMUL_ROW, col).value))
        monthly.append(_num(ws.cell(FACT_TOTAL_ROW, col).value))
    return {"cumul": cumul, "monthly": monthly, "excel": excel.name if excel else None}


def read_savings_balances(closed_month: int, path: Path | None = None) -> dict | None:
    """Ликвидные позиции = старт (кол. B) + сумма месячных потоков.

    Маша накопления: стр. 13/18 (без вычета парковки из формулы августа).
    Саша накопления: стр. 14/19 + Саша инвестиции 17/22 (как в «Разделение_деньги»).
    Наличные: Доллары дома стр. 15/20.
    """
    wb, excel = _wb(path)
    if not wb or FACT_SHEET not in wb.sheetnames:
        return None
    ws = wb[FACT_SHEET]

    def start_plus_flows(bal_row: int, flow_row: int) -> float:
        start = _num(ws.cell(bal_row, 2).value)
        flows = sum(
            _num(ws.cell(flow_row, _fact_col(m)).value)
            for m in range(1, closed_month + 1)
        )
        return start + flows

    masha = start_plus_flows(FACT_MASHA_BAL_ROW, FACT_MASHA_FLOW_ROW)
    sasha_sav = start_plus_flows(FACT_SASHA_BAL_ROW, FACT_SASHA_FLOW_ROW)
    sasha_inv = start_plus_flows(17, 22)  # Саша инвестиции
    cash = start_plus_flows(15, 20)  # Доллары дома
    return {
        "masha": masha,
        "sasha": sasha_sav + sasha_inv,
        "sasha_savings": sasha_sav,
        "sasha_invest": sasha_inv,
        "cash": cash,
        "excel": excel.name if excel else None,
        "method": "start_plus_flows",
    }


def clear_excel_cache() -> None:
    _workbook.cache_clear()
